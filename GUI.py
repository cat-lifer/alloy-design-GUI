# -*- coding: utf-8 -*-
"""
Created on Mon Jul 12 20:41:24 2021

@author: hongyong Han

To: Do or do not. There is no try.

"""
from tkinter import StringVar,IntVar,DoubleVar,Frame,messagebox,TOP,LEFT,Tk,ttk
import microstructure as m
import thermal as t
import numpy as np
import pandas as pd
import openpyxl
import warnings
warnings.filterwarnings("ignore") 

class App:
    def __init__(self,master):
        self.master = master
        self.intWind()
    
    def intWind(self):
       
        self.st=StringVar()        
        self.int1 = IntVar()
        self.int2 = IntVar()
        self.int3 = IntVar()
        self.float1 = DoubleVar()
        self.float2 = DoubleVar()
        self.float3 = DoubleVar()
        self.float4 = DoubleVar()
        self.float5 = DoubleVar()
        
        ## 单个算
        f0 = Frame(self.master)
        f0.pack()
        ttk.Label(f0,text='单个运算',font=('微软雅黑',10,'bold'),width=12.5).pack(side=TOP)
        ttk.Label(f0,text='Ni Al Co Cr Mo Re Ru Ta W',width=25).pack()
        ttk.Entry(f0,textvariable=self.st,width=50).pack(pady=4)
        
        # 创建按钮、绑定方法       
        ttk.Button(f0,text='特征温度',command=self.getT).pack(side=LEFT,padx=25)
        ttk.Button(f0,text='TCP含量',command=self.getTCP).pack(side=LEFT,padx=25)
        ttk.Button(f0,text='组织参数',command=self.getstr).pack(side=LEFT,padx=25)
        
        
        ## 批处理
        f1 = Frame(self.master)
        f1.pack()
        ttk.Label(f1,text='------------------------------------------------------------------------',width=12.5).pack(side=TOP)
        ttk.Label(f1,text='批处理运算',font=('微软雅黑',10,'bold'),width=12.5).pack(side=TOP)
        ttk.Button(f1,text='特征温度',command=self.pgetT).pack(side=LEFT,padx=55)
        ttk.Button(f1,text='TCP含量',command=self.pgetTCP).pack(side=LEFT)
        ttk.Button(f1,text='组织参数',command=self.pgetstr).pack(side=LEFT,padx=55)
        
        ## 合金筛选
        f2 =Frame(self.master)
        f2.pack()
        ttk.Label(f2,text='------------------------------------------------------------------------',width=12.5).pack()
        ttk.Label(f2,text='合金筛选',font=('微软雅黑',10,'bold'),width=12.5).pack()
        ttk.Label(f2,text='固相线 ≥',width=7).pack(side=LEFT,padx=4)        
        ttk.Entry(f2,textvariable=self.int1,width=5).pack(side=LEFT,padx=6)
        
        ttk.Label(f2,text='γ'+'’'+' 溶解温度 ≥',width=10).pack(side=LEFT,padx=4)
        ttk.Entry(f2,textvariable=self.int2,width=5).pack(side=LEFT,padx=6)
        
        ttk.Label(f2,text='热处理窗口 ≥',width=10).pack(side=LEFT,padx=4)
        ttk.Entry(f2,textvariable=self.int3,width=5).pack(side=LEFT,padx=6)
        
        f3 =Frame(self.master)
        f3.pack()
        ttk.Button(f3,text='特征温度筛选',command=self.sgetT).pack(pady=10)
        
        f4 =Frame(self.master)
        f4.pack()
        ttk.Label(f4,text='JMatPro—TCP + 2.11*成分多项式-36.89 ≤ 0',width=38).pack(side=TOP)
        ttk.Button(f4,text='TCP是否析出筛选',command=self.sgetTCP).pack(pady=4)
        
        f5 =Frame(self.master)
        f5.pack()
        ttk.Entry(f5,textvariable=self.float1,width=5).pack(side=LEFT,pady=8)
        ttk.Label(f5,text='≤长宽比≤',width=8).pack(side=LEFT,pady=8)        
        ttk.Entry(f5,textvariable=self.float2,width=5).pack(side=LEFT,pady=8)
        ttk.Label(f5,text='体积分数≥',width=8).pack(side=LEFT,padx=8,pady=8)
        ttk.Entry(f5,textvariable=self.float3,width=5).pack(side=LEFT,pady=8)
        
        f6 =Frame(self.master)
        f6.pack()
        ttk.Entry(f6,textvariable=self.float4,width=5).pack(side=LEFT)
        ttk.Label(f6,text='≤厚度(尺寸)≤',width=11).pack(side=LEFT,pady=5)
        ttk.Entry(f6,textvariable=self.float5,width=5).pack(side=LEFT) 
        
        f7 =Frame(self.master)
        f7.pack()
        ttk.Button(f7,text='组织参数筛选',command=self.sgetstr).pack()
        
        
    ### 单个计算部分的绑定方法  
    def getT(self):
         c = self.st.get()
         c1=c.split()
         cc =np.array(c1).reshape(1, -1)
         
         solves_pred,solidus_pred,liquidus_pred = t.get_thermalparameters(cc)
         results =[solves_pred,solidus_pred,liquidus_pred]
         messagebox.showinfo(title='结果',message=results)
         print(solves_pred,solidus_pred,liquidus_pred)   

    def getTCP(self):            
         c = self.st.get()
         c1=c.split()
         cc =np.array(c1).reshape(1, -1)
         
         TCP = t.get_TCPcontent(cc)
         results = [TCP]
         messagebox.showinfo(title='结果',message=results)
         print(results)
         
    def getstr(self):
        c = self.st.get()
        c1=c.split()
        
        cc =np.array(c1).reshape(1, -1)
        
        set2=cc[:,1:9]
        Feret_Ratio = m.get_FeretRatio(cc)
        gammaprime_thickness = m.get_thickness(cc)
        gammaprime_volume = m.get_volumefraction(set2)
        results =[Feret_Ratio,gammaprime_thickness,gammaprime_volume]
        messagebox.showinfo(title='结果',message=results)
        print(Feret_Ratio,gammaprime_thickness,gammaprime_thickness)
        
    
    ###  批处理部分的绑定方法  
    def pgetT(self):
        cc =np.loadtxt(r'成分集.txt',delimiter='\t')
         
        solves_pred,solidus_pred,liquidus_pred = t.get_thermalparameters(cc)
        heat_window = solidus_pred - solves_pred
        
        total = len(cc)
        data_1 = np.concatenate((cc,solves_pred.reshape(total,1),solidus_pred.reshape(total,1),
                            liquidus_pred.reshape(total,1),heat_window.reshape(total,1)),axis=1)
        
        T = pd.DataFrame(data_1, columns=['Ni','Al','Co','Cr','Mo','Re','Ru','Ta',
                      'W','solves','solidus','liquidus','heatwindow'])
        wb = openpyxl.load_workbook('结果.xlsx')
        writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl')  
        writer.book = wb
        T.to_excel(writer,sheet_name="temperature")
        writer.save()
        results =['已保存到 结果.xlsx/temperature Sheet']
        messagebox.showinfo(title='结果',message=results)
        

    def pgetTCP(self):            
         cc =np.loadtxt(r'成分集.txt',delimiter='\t')
         
         TCP_pred = t.get_TCPcontent(cc)
         
         total = len(cc)
         data_1 = np.concatenate((cc,TCP_pred.reshape(total,1)),axis=1)

         TCP = pd.DataFrame(data_1, columns=['Ni','Al','Co','Cr','Mo','Re','Ru','Ta',
                       'W','TCP(wt.%)'])
         wb = openpyxl.load_workbook('结果.xlsx')
         writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl') 
         writer.book = wb
         TCP.to_excel(writer,sheet_name='tcp')
         writer.save()         
         results = ['已保存到 结果.xlsx/tcp Sheet']
         messagebox.showinfo(title='结果',message=results)  

    def pgetstr(self):
        cc =np.loadtxt(r'成分集.txt',delimiter='\t')
        
        set2=cc[:,1:10]
        Feret_Ratio = m.get_FeretRatio(cc)
        gammaprime_thickness = m.get_thickness(cc)
        gammaprime_volume = m.get_volumefraction(set2)
        
        total = len(cc)
        data_1 = np.concatenate((cc,Feret_Ratio.reshape(total,1),gammaprime_thickness.reshape(total,1),
                             gammaprime_volume.reshape(total,1)),axis=1)
       
        
        structure = pd.DataFrame(data_1, columns=['Ni','Al','Co','Cr','Mo','Re','Ru','Ta',
                       'W','Feret_Ratio','thickness','volumefraction'])
        wb = openpyxl.load_workbook('结果.xlsx')
        writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl') 
        writer.book = wb
        structure.to_excel(writer,sheet_name='microstructure') 
        writer.save()  
        results =['已保存到 结果.xlsx/microstructure Sheet']
        messagebox.showinfo(title='结果',message=results)             
        # print(Feret_Ratio)
        # print(gammaprime_thickness)
        # print(gammaprime_volume)         
      
    ###  筛选部分的绑定方法
    ## 特征温度筛选
    def sgetT(self):  
        data1= pd.read_excel('结果.xlsx','temperature')
        data2= pd.read_excel('结果.xlsx','tcp')
        data3= pd.read_excel('结果.xlsx','microstructure')
        data1['TCP(wt.%)'] =data2['TCP(wt.%)']
        data1['Feret_Ratio'] = data3['Feret_Ratio']
        data1['Feret_Ratio'] = data3['Feret_Ratio']
        data1['thickness']=data3['thickness']
        data1['volumefraction']=data3['volumefraction']
        
        wb = openpyxl.load_workbook('结果.xlsx')
        writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl') 
        writer.book = wb
        data1.to_excel(writer,sheet_name='all_results',index=False) 
        writer.save()
        
        T = data1
        #T = pd.read_excel('结果.xlsx', sheet_name='temperature', index_col=None)
        j=0
        judged=np.zeros((len(T), np.size(T,axis=1)),dtype=float)
        for i in range (len(T)):
             judge_output= T.iloc[i]
             if judge_output['solidus']>=self.int1.get():
                 #print('solidus is ok')
                 if judge_output['solves']>=self.int2.get():  #预测出的值大概比JmatPro计算的低15度
                     #print('solves is ok')
                    if judge_output['heatwindow']>=self.int3.get():
                     #print('heatwindow is ok')                
                       judged[i,:]=judge_output
                       j=j+1
        result1 =['there are %s alloys in search space' %len(T)]
        result2=['there are %s alloys satisfy thermal conditions' %j]             
        messagebox.showinfo(title='结果',message=result1+result2)
        # print("\n")
        # print('there are', len(T) ,'alloys in search space')  
        # print('there are', j, 'alloys satisfy thermal conditions')    
        # print("\n")
        
        thermal_selected=np.zeros((j, np.size(judged,axis=1)),dtype=float)
        l=0
        for k in range (len(judged)): 
            if  judged[k,:].any()!=0: 
                thermal_selected[l,:]=judged[k,:]    
                l+=1              
        selected_1 = pd.DataFrame(thermal_selected, columns=['index','Ni','Al','Co','Cr','Mo','Re','Ru','Ta',
                      'W','solvus','solidus','liquidus','heatwindow','TCP(wt.%)'
                      ,'Feret_Ratio','thickness','volumefraction'])
        wb = openpyxl.load_workbook('结果.xlsx')
        writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl')  
        writer.book = wb
        selected_1.to_excel(writer,sheet_name="after_filter_T",index=False)
        writer.save()
    
    ## TCP筛选    
    def sgetTCP(self):
        selected_2 = pd.read_excel('结果.xlsx', sheet_name='after_filter_T', index_col=None)
        JmatPro_TCP = selected_2['TCP(wt.%)']
        # 计算成分多项式
        Ni = selected_2['Ni']/58.69
        Al = selected_2['Al']/26.98
        Co = selected_2['Co']/58.93
        Cr = selected_2['Cr']/52
        Mo = selected_2['Mo']/95.94
        Re = selected_2['Re']/186.2
        Ru = selected_2['Ru']/101.1
        Ta = selected_2['Ta']/180.9
        W  = selected_2['W']/183.8

        total = Ni+Al+Co+Cr+Mo+Re+Ru+Ta+W
        #Atom_Ni = (Ni/total)*100
        Atom_Al = (Al/total)*100
        #Atom_Co = (Co/total)*100
        Atom_Cr = (Cr/total)*100  
        Atom_Mo = (Mo/total)*100
        Atom_Re = (Re/total)*100
        Atom_Ru = (Ru/total)*100  
        Atom_Ta = (Ta/total)*100
        Atom_W  = (W/total)*100
        
        composition_polynomial = (Atom_Ta+1.5*Atom_Re+0.75*(Atom_Cr+Atom_Mo)+
                         0.5*(Atom_Al+Atom_W)- 0.4*Atom_Ru)
        
        # 计算TCP判式
        TCP_judgement = m.TCP_judge(JmatPro_TCP, composition_polynomial)
        TCP_judgement = np.array(TCP_judgement).reshape(len(selected_2),1)
        
        selected_2['TCP_judgement']=TCP_judgement
        
        j=0
        judged=np.zeros((len(selected_2), np.size(selected_2,axis=1)),dtype=float)
        for i in range (len(selected_2)):
            judge_output= selected_2.iloc[i]
            if judge_output['TCP_judgement']<=0:
                    judged[i,:]=judge_output
                    j=j+1
        
        result1 =['there are %s alloys in search space' %len(selected_2)]
        result2=['there are %s alloys satisfy TCP conditions' %j]             
        messagebox.showinfo(title='结果',message=result1+result2)
        # print("\n")
        # print('there are', len(selected_2) ,'alloys in search space')  
        # print('there are', j, 'alloys satisfy TCP conditions')    
        # print("\n")

        TCP_selected=np.zeros((j, np.size(judged,axis=1)),dtype=float)
        l=0
        for k in range (len(judged)): 
           if  judged[k,:].any()!=0: 
               TCP_selected[l,:]=judged[k,:]    
               l+=1          
        selected_3 = pd.DataFrame(TCP_selected, columns=['index','Ni','Al','Co','Cr','Mo','Re','Ru','Ta',
                      'W','solvus','solidus','liquidus','heatwindow','TCP(wt.%)'
                      ,'Feret_Ratio','thickness','volumefraction','TCP_judgement'])
        wb = openpyxl.load_workbook('结果.xlsx')
        writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl')  
        writer.book = wb
        selected_3.to_excel(writer,sheet_name="after_filter_TCP",index=False)
        writer.save()        
        
    ## 组织参数筛选    
    def sgetstr(self):
        Str = pd.read_excel('结果.xlsx', sheet_name='after_filter_TCP', index_col=None)
        j=0
        judged=np.zeros((len(Str), np.size(Str,axis=1)),dtype=float)
        for i in range (len(Str)):
             judge_output= Str.iloc[i]
             if (judge_output['Feret_Ratio']>=self.float1.get() and judge_output['Feret_Ratio']<=self.float2.get()):
                 #print('solidus is ok')
                 if (judge_output['thickness']>=self.float4.get() and judge_output['thickness']<=self.float5.get()):  
                     #print('solves is ok')
                    if judge_output['volumefraction']>=self.float3.get():
                     #print('heatwindow is ok')                
                       judged[i,:]=judge_output
                       j=j+1
        
        result1 =['there are %s alloys in search space' %len(Str)]
        result2=['there are %s alloys satisfy microstructure conditions' %j]             
        messagebox.showinfo(title='结果',message=result1+result2)
        # print("\n")
        # print('there are', len(Str) ,'alloys in search space')  
        # print('there are', j, 'alloys satisfy microstructure conditions')    
        # print("\n")
        
        str_selected=np.zeros((j, np.size(judged,axis=1)),dtype=float)
        l=0
        for k in range (len(judged)): 
           if  judged[k,:].any()!=0: 
               str_selected[l,:]=judged[k,:]    
               l+=1          
        selected_2 = pd.DataFrame(str_selected, columns=['index','Ni','Al','Co','Cr','Mo','Re','Ru','Ta',
                      'W','solvus','solidus','liquidus','heatwindow','TCP(wt.%)'
                      ,'Feret_Ratio','thickness','volumefraction','TCP_judgement'])
        wb = openpyxl.load_workbook('结果.xlsx')
        writer = pd.ExcelWriter('结果.xlsx',engine='openpyxl')  
        writer.book = wb
        selected_2.to_excel(writer,sheet_name="after_filter_structure",index=False)
        writer.save() 
        
    
           
from icon import img
import os
import base64       

if __name__ == '__main__':
    root = Tk()
    
    tmp = open("tmp.ico", "wb+")
    tmp.write(base64.b64decode(img))
    tmp.close()
    root.iconbitmap('tmp.ico')  # 加图标
    os.remove("tmp.ico")  #
    
    root.geometry('650x580')
    root.title("工具箱")
    App(root)
    root.mainloop()




           